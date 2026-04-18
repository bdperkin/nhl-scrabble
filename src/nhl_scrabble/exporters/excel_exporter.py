"""Excel export functionality for NHL Scrabble reports."""

from __future__ import annotations

from pathlib import Path  # noqa: TC003
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from nhl_scrabble.models.player import PlayerScore
from nhl_scrabble.models.team import TeamScore


class ExcelExporter:
    """Export NHL Scrabble data to Excel format.

    Provides methods to export various report types to Excel files with
    multiple sheets, formatting, and styling for better readability.

    Requires openpyxl to be installed. Install with:
        pip install nhl-scrabble[export]

    Examples:
        >>> exporter = ExcelExporter()
        >>> exporter.export_full_report(data, Path("report.xlsx"))
        >>> exporter.export_team_scores(teams, Path("teams.xlsx"))
    """

    def __init__(self) -> None:
        """Initialize Excel exporter.

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install nhl-scrabble[export]"
            )

    def _format_header_row(self, ws: Any, row: int = 1) -> None:
        """Apply formatting to header row.

        Args:
            ws: Worksheet object
            row: Row number for header (default: 1)
        """
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def _auto_adjust_columns(self, ws: Any) -> None:
        """Auto-adjust column widths based on content.

        Args:
            ws: Worksheet object
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (AttributeError, TypeError):
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_team_scores(
        self, teams: dict[str, TeamScore] | list[TeamScore], output: Path
    ) -> None:
        """Export team scores to Excel file.

        Args:
            teams: Dictionary or list of TeamScore objects
            output: Output file path

        Raises:
            OSError: If file cannot be written
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Team Scores"

        # Headers
        ws.append(
            [
                "Team",
                "Division",
                "Conference",
                "Total Score",
                "Player Count",
                "Average Score",
            ]
        )
        self._format_header_row(ws)

        # Convert dict to list if necessary
        team_list = list(teams.values()) if isinstance(teams, dict) else teams

        # Sort by total score descending
        sorted_teams = sorted(team_list, key=lambda t: (t.total, t.avg_per_player), reverse=True)

        # Data rows
        for team in sorted_teams:
            ws.append(
                [
                    team.abbrev,
                    team.division,
                    team.conference,
                    team.total,
                    team.player_count,
                    round(team.avg_per_player, 2),
                ]
            )

        self._auto_adjust_columns(ws)
        wb.save(output)

    def export_player_scores(self, players: list[PlayerScore], output: Path) -> None:
        """Export player scores to Excel file.

        Args:
            players: List of PlayerScore objects
            output: Output file path

        Raises:
            OSError: If file cannot be written
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Player Scores"

        # Headers
        ws.append(
            [
                "Player Name",
                "Team",
                "Division",
                "Conference",
                "First Name Score",
                "Last Name Score",
                "Total Score",
            ]
        )
        self._format_header_row(ws)

        # Sort by score descending
        sorted_players = sorted(players, key=lambda p: p.full_score, reverse=True)

        # Data rows
        for player in sorted_players:
            ws.append(
                [
                    player.full_name,
                    player.team,
                    player.division,
                    player.conference,
                    player.first_score,
                    player.last_score,
                    player.full_score,
                ]
            )

        self._auto_adjust_columns(ws)
        wb.save(output)

    def export_full_report(  # noqa: C901, PLR0912
        self,
        team_scores: dict[str, TeamScore],
        all_players: list[PlayerScore],
        division_standings: dict[str, Any],
        conference_standings: dict[str, Any],
        playoff_standings: dict[str, Any],
        output: Path,
        sheets: list[str] | None = None,
    ) -> None:
        """Export full report with multiple sheets to Excel file.

        Args:
            team_scores: Dictionary of team scores
            all_players: List of all players
            division_standings: Dictionary of division standings
            conference_standings: Dictionary of conference standings
            playoff_standings: Dictionary of playoff standings
            output: Output file path
            sheets: Optional list of sheets to include (default: all)
                Options: 'teams', 'players', 'divisions', 'conferences', 'playoffs'

        Raises:
            OSError: If file cannot be written
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Determine which sheets to create
        if sheets is None:
            sheets = ["teams", "players", "divisions", "conferences", "playoffs"]

        # Teams sheet
        if "teams" in sheets:
            ws_teams = wb.create_sheet("Teams")
            ws_teams.append(
                [
                    "Team",
                    "Division",
                    "Conference",
                    "Total Score",
                    "Player Count",
                    "Average Score",
                ]
            )
            self._format_header_row(ws_teams)

            sorted_teams = sorted(
                team_scores.values(), key=lambda t: (t.total, t.avg_per_player), reverse=True
            )
            for team in sorted_teams:
                ws_teams.append(
                    [
                        team.abbrev,
                        team.division,
                        team.conference,
                        team.total,
                        team.player_count,
                        round(team.avg_per_player, 2),
                    ]
                )
            self._auto_adjust_columns(ws_teams)

        # Players sheet
        if "players" in sheets:
            ws_players = wb.create_sheet("Players")
            ws_players.append(
                [
                    "Player Name",
                    "Team",
                    "Division",
                    "Conference",
                    "First Name Score",
                    "Last Name Score",
                    "Total Score",
                ]
            )
            self._format_header_row(ws_players)

            sorted_players = sorted(all_players, key=lambda p: p.full_score, reverse=True)
            for player in sorted_players:
                ws_players.append(
                    [
                        player.full_name,
                        player.team,
                        player.division,
                        player.conference,
                        player.first_score,
                        player.last_score,
                        player.full_score,
                    ]
                )
            self._auto_adjust_columns(ws_players)

        # Divisions sheet
        if "divisions" in sheets:
            ws_divisions = wb.create_sheet("Divisions")
            ws_divisions.append(
                [
                    "Division",
                    "Team",
                    "Total Score",
                    "Player Count",
                    "Average Score",
                ]
            )
            self._format_header_row(ws_divisions)

            for division_name in sorted(division_standings.keys()):
                standing = division_standings[division_name]
                for team in standing.teams:
                    ws_divisions.append(
                        [
                            division_name,
                            team.abbrev,
                            team.total,
                            team.player_count,
                            round(team.avg_per_player, 2),
                        ]
                    )
            self._auto_adjust_columns(ws_divisions)

        # Conferences sheet
        if "conferences" in sheets:
            ws_conferences = wb.create_sheet("Conferences")
            ws_conferences.append(
                [
                    "Conference",
                    "Team",
                    "Division",
                    "Total Score",
                    "Player Count",
                    "Average Score",
                ]
            )
            self._format_header_row(ws_conferences)

            for conference_name in sorted(conference_standings.keys()):
                standing = conference_standings[conference_name]
                for team in standing.teams:
                    ws_conferences.append(
                        [
                            conference_name,
                            team.abbrev,
                            team.division,
                            team.total,
                            team.player_count,
                            round(team.avg_per_player, 2),
                        ]
                    )
            self._auto_adjust_columns(ws_conferences)

        # Playoffs sheet
        if "playoffs" in sheets:
            ws_playoffs = wb.create_sheet("Playoffs")
            ws_playoffs.append(
                [
                    "Conference",
                    "Seed",
                    "Team",
                    "Division",
                    "Total Score",
                    "Average Score",
                    "Status",
                ]
            )
            self._format_header_row(ws_playoffs)

            for conference_name in sorted(playoff_standings.keys()):
                teams_list = playoff_standings[conference_name]
                for seed, team in enumerate(teams_list, start=1):
                    ws_playoffs.append(
                        [
                            conference_name,
                            seed,
                            team.abbrev,
                            team.division,
                            team.total,
                            round(team.avg, 2),
                            team.status_indicator,
                        ]
                    )
            self._auto_adjust_columns(ws_playoffs)

        wb.save(output)
