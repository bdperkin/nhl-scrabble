"""Tests for Excel exporter."""

from pathlib import Path
from typing import Any

import pytest

from nhl_scrabble.exporters.excel_exporter import OPENPYXL_AVAILABLE, ExcelExporter
from nhl_scrabble.models.player import PlayerScore
from nhl_scrabble.models.standings import ConferenceStandings, DivisionStandings, PlayoffTeam
from nhl_scrabble.models.team import TeamScore

if OPENPYXL_AVAILABLE:
    from openpyxl import load_workbook


pytestmark = pytest.mark.skipif(
    not OPENPYXL_AVAILABLE,
    reason="openpyxl not installed (optional dependency)",
)


@pytest.fixture
def excel_exporter() -> ExcelExporter:
    """Create Excel exporter instance."""
    return ExcelExporter()


@pytest.fixture
def sample_teams() -> dict[str, TeamScore]:
    """Create sample team scores."""
    players_tor = [
        PlayerScore(
            first_name="John",
            last_name="Tavares",
            full_name="John Tavares",
            first_score=10,
            last_score=15,
            full_score=25,
            team="TOR",
            division="Atlantic",
            conference="Eastern",
        ),
        PlayerScore(
            first_name="Auston",
            last_name="Matthews",
            full_name="Auston Matthews",
            first_score=18,
            last_score=20,
            full_score=38,
            team="TOR",
            division="Atlantic",
            conference="Eastern",
        ),
    ]

    players_mtl = [
        PlayerScore(
            first_name="Nick",
            last_name="Suzuki",
            full_name="Nick Suzuki",
            first_score=8,
            last_score=25,
            full_score=33,
            team="MTL",
            division="Atlantic",
            conference="Eastern",
        ),
    ]

    return {
        "TOR": TeamScore(
            abbrev="TOR",
            total=63,
            players=players_tor,
            division="Atlantic",
            conference="Eastern",
        ),
        "MTL": TeamScore(
            abbrev="MTL",
            total=33,
            players=players_mtl,
            division="Atlantic",
            conference="Eastern",
        ),
    }


@pytest.fixture
def sample_players() -> list[PlayerScore]:
    """Create sample player scores."""
    return [
        PlayerScore(
            first_name="John",
            last_name="Tavares",
            full_name="John Tavares",
            first_score=10,
            last_score=15,
            full_score=25,
            team="TOR",
            division="Atlantic",
            conference="Eastern",
        ),
        PlayerScore(
            first_name="Auston",
            last_name="Matthews",
            full_name="Auston Matthews",
            first_score=18,
            last_score=20,
            full_score=38,
            team="TOR",
            division="Atlantic",
            conference="Eastern",
        ),
    ]


def test_excel_exporter_initialization() -> None:
    """Test Excel exporter initializes correctly when openpyxl is available."""
    exporter = ExcelExporter()
    assert exporter is not None


def test_export_team_scores_dict(
    excel_exporter: ExcelExporter,
    sample_teams: dict[str, TeamScore],
    tmp_path: Path,
) -> None:
    """Test exporting team scores from dictionary to Excel."""
    output = tmp_path / "teams.xlsx"

    excel_exporter.export_team_scores(sample_teams, output)

    assert output.exists()

    # Load and verify content
    wb = load_workbook(output)
    ws = wb.active
    assert ws.title == "Team Scores"

    # Check header
    headers = [cell.value for cell in ws[1]]
    assert headers == [
        "Team",
        "Division",
        "Conference",
        "Total Score",
        "Player Count",
        "Average Score",
    ]

    # Check data (TOR should be first with higher score)
    row2 = [cell.value for cell in ws[2]]
    assert row2[0] == "TOR"
    assert row2[3] == 63
    assert row2[4] == 2
    assert row2[5] == 31.5


def test_export_team_scores_list(
    excel_exporter: ExcelExporter,
    sample_teams: dict[str, TeamScore],
    tmp_path: Path,
) -> None:
    """Test exporting team scores from list to Excel."""
    output = tmp_path / "teams.xlsx"
    teams_list = list(sample_teams.values())

    excel_exporter.export_team_scores(teams_list, output)

    assert output.exists()
    wb = load_workbook(output)
    ws = wb.active
    assert ws.title == "Team Scores"


def test_export_player_scores(
    excel_exporter: ExcelExporter,
    sample_players: list[PlayerScore],
    tmp_path: Path,
) -> None:
    """Test exporting player scores to Excel."""
    output = tmp_path / "players.xlsx"

    excel_exporter.export_player_scores(sample_players, output)

    assert output.exists()

    # Load and verify content
    wb = load_workbook(output)
    ws = wb.active
    assert ws.title == "Player Scores"

    # Check header
    headers = [cell.value for cell in ws[1]]
    assert headers == [
        "Player Name",
        "Team",
        "Division",
        "Conference",
        "First Name Score",
        "Last Name Score",
        "Total Score",
    ]

    # Check first player (sorted by score, Auston Matthews has 38)
    row2 = [cell.value for cell in ws[2]]
    assert row2[0] == "Auston Matthews"
    assert row2[6] == 38


def test_export_full_report_all_sheets(
    excel_exporter: ExcelExporter,
    sample_teams: dict[str, TeamScore],
    tmp_path: Path,
) -> None:
    """Test exporting full report with all sheets."""
    output = tmp_path / "full_report.xlsx"

    # Create sample data
    all_players = []
    for team in sample_teams.values():
        all_players.extend(team.players)

    # Create standings objects with team abbreviations (as per actual model)
    team_abbrevs = list(sample_teams.keys())
    total_score = sum(t.total for t in sample_teams.values())
    total_players = sum(t.player_count for t in sample_teams.values())

    division_standings = {
        "Atlantic": DivisionStandings(
            name="Atlantic",
            total=total_score,
            teams=team_abbrevs,
            player_count=total_players,
            avg_per_team=total_score / len(team_abbrevs),
        ),
    }

    conference_standings = {
        "Eastern": ConferenceStandings(
            name="Eastern",
            total=total_score,
            teams=team_abbrevs,
            player_count=total_players,
            avg_per_team=total_score / len(team_abbrevs),
        ),
    }

    # Create playoff standings using the PlayoffTeam model
    playoff_standings = {
        "Eastern": [
            PlayoffTeam(
                abbrev="TOR",
                total=63,
                players=2,
                avg=31.5,
                conference="Eastern",
                division="Atlantic",
                status_indicator="y",
            ),
        ],
    }

    excel_exporter.export_full_report(
        team_scores=sample_teams,
        all_players=all_players,
        division_standings=division_standings,
        conference_standings=conference_standings,
        playoff_standings=playoff_standings,
        output=output,
    )

    assert output.exists()

    # Load and verify sheets
    wb = load_workbook(output)
    sheet_names = wb.sheetnames

    assert "Teams" in sheet_names
    assert "Players" in sheet_names
    assert "Divisions" in sheet_names
    assert "Conferences" in sheet_names
    assert "Playoffs" in sheet_names


def test_export_full_report_selected_sheets(
    excel_exporter: ExcelExporter,
    sample_teams: dict[str, TeamScore],
    tmp_path: Path,
) -> None:
    """Test exporting full report with selected sheets only."""
    output = tmp_path / "selected_report.xlsx"

    # Create minimal data
    all_players = []
    for team in sample_teams.values():
        all_players.extend(team.players)

    division_standings: dict[str, Any] = {}
    conference_standings: dict[str, Any] = {}
    playoff_standings: dict[str, Any] = {}

    excel_exporter.export_full_report(
        team_scores=sample_teams,
        all_players=all_players,
        division_standings=division_standings,
        conference_standings=conference_standings,
        playoff_standings=playoff_standings,
        output=output,
        sheets=["teams", "players"],
    )

    assert output.exists()

    # Load and verify only selected sheets exist
    wb = load_workbook(output)
    sheet_names = wb.sheetnames

    assert "Teams" in sheet_names
    assert "Players" in sheet_names
    assert "Divisions" not in sheet_names
    assert "Conferences" not in sheet_names
    assert "Playoffs" not in sheet_names


def test_excel_header_formatting(
    excel_exporter: ExcelExporter,
    sample_teams: dict[str, TeamScore],
    tmp_path: Path,
) -> None:
    """Test that Excel headers have proper formatting."""
    output = tmp_path / "teams.xlsx"

    excel_exporter.export_team_scores(sample_teams, output)

    wb = load_workbook(output)
    ws = wb.active

    # Check header cell formatting
    first_cell = ws["A1"]
    assert first_cell.font.bold is True
    # Check fill color (may have different formats, just check it's set)
    assert first_cell.fill.start_color.rgb is not None
    assert "366092" in first_cell.fill.start_color.rgb


def test_excel_column_widths(
    excel_exporter: ExcelExporter,
    sample_players: list[PlayerScore],
    tmp_path: Path,
) -> None:
    """Test that Excel columns are auto-adjusted."""
    output = tmp_path / "players.xlsx"

    excel_exporter.export_player_scores(sample_players, output)

    wb = load_workbook(output)
    ws = wb.active

    # Check that columns have been adjusted (not default width)
    for col_letter in ["A", "B", "C"]:
        width = ws.column_dimensions[col_letter].width
        assert width > 0  # Should be adjusted


def test_excel_encoding(excel_exporter: ExcelExporter, tmp_path: Path) -> None:
    """Test Excel handles UTF-8 encoding correctly."""
    output = tmp_path / "players.xlsx"

    # Create player with accented name
    players = [
        PlayerScore(
            first_name="François",
            last_name="Beauchemin",
            full_name="François Beauchemin",
            first_score=20,
            last_score=25,
            full_score=45,
            team="MTL",
            division="Atlantic",
            conference="Eastern",
        ),
    ]

    excel_exporter.export_player_scores(players, output)

    assert output.exists()

    wb = load_workbook(output)
    ws = wb.active
    player_name = ws["A2"].value
    assert player_name == "François Beauchemin"


def test_excel_sorting(excel_exporter: ExcelExporter, tmp_path: Path) -> None:
    """Test that Excel exports are properly sorted."""
    output = tmp_path / "teams.xlsx"

    teams = [
        TeamScore(
            abbrev="MTL",
            total=30,
            players=[],
            division="Atlantic",
            conference="Eastern",
        ),
        TeamScore(
            abbrev="TOR",
            total=50,
            players=[],
            division="Atlantic",
            conference="Eastern",
        ),
        TeamScore(
            abbrev="BOS",
            total=40,
            players=[],
            division="Atlantic",
            conference="Eastern",
        ),
    ]

    excel_exporter.export_team_scores(teams, output)

    wb = load_workbook(output)
    ws = wb.active

    # Check sorting (descending by total score)
    assert ws["A2"].value == "TOR"  # 50 points
    assert ws["A3"].value == "BOS"  # 40 points
    assert ws["A4"].value == "MTL"  # 30 points
